import openpyxl
import unicodedata


def convert_to_dict(file_name):
    data = openpyxl.load_workbook(
        f"/Users/gimdongmin/Desktop/memorize_english/datas/{file_name}.xlsx"
    )
    ws = data.active
    words = ws["B"] + ws["D"]
    title = ws.cell(row=1, column=1).value
    all_value = []
    for cell in words:
        all_value.append(unicodedata.normalize("NFKD", str(cell.value)))

    return {f"{title}": all_value}

\
