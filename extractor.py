import openpyxl
import unicodedata
import os
import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()
from words import models


base_path = "/Users/gimdongmin/Desktop/memorize_english/datas/words/"
file_name = "new_기본영어단어1000개.xlsx"
full_path = f"{base_path}/{file_name}"

def convert_data():
    data = openpyxl.load_workbook(
        full_path

    )
    ws = data.active
    words = ws["B"] + ws["E"]
    meaning = ws["C"] + ws["F"]
    title = ws.cell(row=1, column=1).value
    all_value = []
    for i in range(len(words)):
        if str(words[i].value) != "None" or str(meaning[i].value) != "None":
            all_value.append(
                [
                    unicodedata.normalize("NFKD", str(words[i].value)),
                    unicodedata.normalize("NFKD", str(meaning[i].value)),
                ]
            )
    models.Word.objects.create(title=title, content=all_value)


convert_data()
